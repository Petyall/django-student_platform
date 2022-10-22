from openpyxl import load_workbook
from pathlib import Path


def parser(request):
    group = request.user.group
    if len(group) == 5:
        a = {}
        keys = 'A', 'B', 'C', 'D', 'E', 'F'
        group = request.user.group
        BASE_DIR = Path(__file__).resolve().parent.parent
        base_str = str(BASE_DIR)
        filepath = base_str + "\\schedule\\" + 'schedule.xlsx'
        book = load_workbook(filename=filepath)
        sheet = book[group]
        i = 0

        for item in book[group]:
            i = i + 1

        for key in keys:
            a.setdefault(key,[])
            for j in range(1, i+1):
                a[key].append(sheet[key + str(j)].value)

    else:
        last_name = request.user.last_name
        name = request.user.first_name
        FI = last_name + ' ' + name
        BASE_DIR = Path(__file__).resolve().parent.parent
        base_str = str(BASE_DIR)
        filepath = base_str + "\\schedule\\" + 'schedule.xlsx'
        book = load_workbook(filename=filepath)
        try:
            book[FI]
            sheet = book[FI]
            i = 0
            for item in book[FI]:
                i = i + 1
            a = {}
            keys = 'A', 'B', 'C', 'D', 'E', 'F'
            for key in keys:
                a.setdefault(key,[])
                for j in range(1, i+1):
                    a[key].append(sheet[key + str(j)].value)   
        except:
            a = 'Расписания нет' 
   
    context = {
        'text': a
    }

    return (context)
